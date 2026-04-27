#!/usr/bin/env python3
"""
GT38 生产排程
- 排产周期: 4/1-4/4 (小批) + 4/5-4/8 (大头)
- 出货节点: 4月4日前 + 4月8日前需生产完
- 验证页: 库存 → 出货 → 余额 → 缺口 → 生产分配
"""

import math
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pathlib import Path

SRC = Path(
    "/Users/kinghinchan/Library/Containers/com.tencent.xinWeChat/Data/Documents/"
    "xwechat_files/wxid_gtzu9fn6a66j12_7b19/temp/drag/4.12前.xls"
)
OUT = Path(__file__).parent.parent / "reports" / "刀刀电子_GT38生产排程.xlsx"

DAILY_CAP = 5000
# 两个排产周期各4天
P1_DAYS = 4  # 4/1-4/4
P2_DAYS = 4  # 4/5-4/8
P1_CAP = DAILY_CAP * P1_DAYS  # 20,000
P2_CAP = DAILY_CAP * P2_DAYS  # 20,000

# Labels
P1 = "4/1-4/4"
P2 = "4/5-4/8"


def c(v):
    try:
        return float(v) if pd.notna(v) else 0
    except (ValueError, TypeError):
        return 0


def read_source():
    df = pd.read_excel(SRC, header=None)
    rows = []
    for i in range(3, len(df) - 1):
        r = df.iloc[i]
        if pd.isna(r[0]) and isinstance(r[10], str):
            continue

        rows.append({
            "产品名称": r[0],
            "客户型号": r[1],
            "包装单位": c(r[2]),
            "4月计划出货": c(r[3]),
            "深圳库存_包装": c(r[4]),
            "深圳库存": c(r[5]),
            "河南库存_包装": c(r[6]),
            "河南库存": c(r[7]),
            "总库存": c(r[8]),
            "需要生产": c(r[9]),
            # 出货节点 (from source sheet)
            "4.4_包装数": c(r[10]),
            "4.4_出货个": c(r[11]),
            "4.12_包装数": c(r[13]),
            "4.12_出货个": c(r[14]),
            "4.19_包装数": c(r[16]),
            "4.19_出货个": c(r[17]),
        })

    # Waterfall
    for r in rows:
        inv = r["总库存"]
        w1 = r["4.4_出货个"]
        w2 = r["4.12_出货个"]
        w3 = r["4.19_出货个"]
        r["4.4后余额"] = inv - w1
        r["4.12后余额"] = inv - w1 - w2
        r["4.19后余额"] = inv - w1 - w2 - w3
        r["4.4缺口"] = max(-r["4.4后余额"], 0)
        r["4.12缺口"] = max(-(inv - w1 - w2), 0)
        r["总出货"] = w1 + w2 + w3
        r["总缺口"] = max(r["总出货"] - inv, 0)

    return rows


def allocate(rows):
    """
    4/1-4/4: 小批, 只做有4/4出货缺口的紧急SKU, 上限 P1_CAP
    4/5-4/8: 大头, 其余全部
    """
    # 4/1-4/4 上限: 2天产量 (10,000), 留空间给4/5-4/8
    W1_TARGET = DAILY_CAP * 2  # 10,000

    for r in rows:
        need = max(r["需要生产"], 0)
        r["总生产"] = need
        r[f"生产_{P1}"] = 0
        r[f"生产_{P2}"] = 0

    prod = [r for r in rows if r["总生产"] > 0]
    prod.sort(key=lambda r: -r["4.4缺口"])

    # Fill P1 with urgent items
    w1_used = 0
    for r in prod:
        gap = r["4.4缺口"]
        if gap <= 0 or w1_used >= W1_TARGET:
            continue
        alloc = min(gap, r["总生产"], W1_TARGET - w1_used)
        r[f"生产_{P1}"] = alloc
        w1_used += alloc

    # Rest goes to P2
    for r in rows:
        if r["总生产"] > 0:
            r[f"生产_{P2}"] = r["总生产"] - r[f"生产_{P1}"]

    # Overflow: if P2 > P2_CAP, push back to P1
    tw2 = sum(r[f"生产_{P2}"] for r in rows)
    tw1 = sum(r[f"生产_{P1}"] for r in rows)
    if tw2 > P2_CAP:
        spare = P1_CAP - tw1
        overflow = tw2 - P2_CAP
        shift = min(spare, overflow)
        if shift > 0:
            print(f"  {P2}超产能 {overflow:,.0f}, 回填{P1} {shift:,.0f}")
            items = sorted([r for r in rows if r[f"生产_{P2}"] > 0], key=lambda r: -r[f"生产_{P2}"])
            rem = shift
            for r in items:
                if rem <= 0:
                    break
                mv = min(r[f"生产_{P2}"], rem)
                r[f"生产_{P1}"] += mv
                r[f"生产_{P2}"] -= mv
                rem -= mv

    tw1 = sum(r[f"生产_{P1}"] for r in rows)
    tw2 = sum(r[f"生产_{P2}"] for r in rows)
    d1 = math.ceil(tw1 / DAILY_CAP) if tw1 else 0
    d2 = math.ceil(tw2 / DAILY_CAP) if tw2 else 0
    print(f"  {P1}: {tw1:>8,.0f} ({d1}天/{P1_DAYS}天)")
    print(f"  {P2}: {tw2:>8,.0f} ({d2}天/{P2_DAYS}天)")
    print(f"  合计:     {tw1+tw2:>8,.0f} ({d1+d2}天)")
    return rows


def write_excel(rows):
    wb = openpyxl.Workbook()
    thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    wf = Font(color="FFFFFF", bold=True, size=10)
    bf = Font(bold=True, size=11)
    rf = Font(color="FF0000", bold=True)

    F = {
        "blue": PatternFill("solid", fgColor="4472C4"),
        "dblue": PatternFill("solid", fgColor="2F5496"),
        "gray": PatternFill("solid", fgColor="A5A5A5"),
        "green": PatternFill("solid", fgColor="548235"),
        "orange": PatternFill("solid", fgColor="ED7D31"),
        "lg": PatternFill("solid", fgColor="E2EFDA"),
        "lb": PatternFill("solid", fgColor="D6E4F0"),
        "lo": PatternFill("solid", fgColor="FCE4D6"),
        "lr": PatternFill("solid", fgColor="FFC7CE"),
        "banner": PatternFill("solid", fgColor="1F4E79"),
    }

    tw1 = sum(r[f"生产_{P1}"] for r in rows)
    tw2 = sum(r[f"生产_{P2}"] for r in rows)
    ta = tw1 + tw2

    # ═══════════════════════════════════════════
    # Sheet 1: 生产排程
    # ═══════════════════════════════════════════
    ws = wb.active
    ws.title = "生产排程"

    ws.merge_cells("A1:R1")
    cell = ws.cell(row=1, column=1,
        value=(f"GT38 生产排程  |  日产能 {DAILY_CAP:,}  |  "
               f"{P1} → {tw1:,.0f}  |  {P2} → {tw2:,.0f}  |  总 {ta:,.0f}"))
    cell.font = Font(bold=True, size=13, color="FFFFFF")
    cell.fill = F["banner"]
    cell.alignment = Alignment(horizontal="center", vertical="center")

    hdrs = [
        ("产品名称", 40, "blue"),
        ("客户型号", 18, "blue"),
        ("包装\n单位", 7, "blue"),
        ("4月计划\n出货", 10, "gray"),
        ("深圳库存", 10, "gray"),
        ("河南库存", 10, "gray"),
        ("总库存", 10, "blue"),
        ("4/4前\n出货", 10, "gray"),
        ("4/4后\n余额", 10, "gray"),
        ("4/12前\n出货", 10, "gray"),
        ("4/12后\n余额", 10, "gray"),
        ("4/19前\n出货", 10, "gray"),
        ("需要\n生产", 10, "orange"),
        (f"生产\n{P1}", 12, "green"),
        (f"生产\n{P2}", 12, "dblue"),
        (f"4/4后余额\n(含{P1}生产)", 14, "green"),
        (f"4/12后余额\n(含全部生产)", 14, "dblue"),
        ("验证\n✓/✗", 7, "blue"),
    ]

    for col, (h, w, fk) in enumerate(hdrs, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = wf
        cell.fill = F[fk]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    prod = sorted([r for r in rows if r["总生产"] > 0], key=lambda r: -r["总生产"])
    noprod = [r for r in rows if r["总生产"] <= 0]

    def emit(ws, ri, r):
        inv = r["总库存"]
        w1_ship = r["4.4_出货个"]
        w2_ship = r["4.12_出货个"]
        w3_ship = r["4.19_出货个"]
        p1 = r[f"生产_{P1}"]
        p2 = r[f"生产_{P2}"]

        bal_after_w1 = inv + p1 - w1_ship
        bal_after_w2 = inv + p1 + p2 - w1_ship - w2_ship
        ok = (r["4月计划出货"] <= inv + p1 + p2) or r["总生产"] <= 0

        vals = [
            r["产品名称"], r["客户型号"], r["包装单位"],
            r["4月计划出货"], r["深圳库存"], r["河南库存"], inv,
            w1_ship, r["4.4后余额"],
            w2_ship, r["4.12后余额"],
            w3_ship,
            r["需要生产"], p1, p2,
            bal_after_w1, bal_after_w2,
            "✓" if ok else "✗",
        ]

        for col, v in enumerate(vals, 1):
            if v == 0 and col >= 4 and col != 7:
                continue
            cell = ws.cell(row=ri, column=col, value=v)
            cell.border = thin
            if col >= 3 and col != 18:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
            elif col == 18:
                cell.alignment = Alignment(horizontal="center")

            if col == 7:
                cell.font = bf
            elif col == 9 and r["4.4后余额"] < 0:
                cell.fill = F["lr"]; cell.font = rf
            elif col == 11 and r["4.12后余额"] < 0:
                cell.fill = F["lr"]; cell.font = rf
            elif col == 13 and r["需要生产"] > 0:
                cell.fill = F["lo"]
            elif col == 14 and p1 > 0:
                cell.fill = F["lg"]; cell.font = bf
            elif col == 15 and p2 > 0:
                cell.fill = F["lb"]; cell.font = bf
            elif col == 16:
                cell.fill = F["lg"] if bal_after_w1 >= 0 else F["lr"]
                if bal_after_w1 < 0: cell.font = rf
            elif col == 17:
                cell.fill = F["lb"] if bal_after_w2 >= 0 else F["lr"]
                if bal_after_w2 < 0: cell.font = rf
            elif col == 18:
                cell.font = Font(color="548235", bold=True) if ok else rf

    ri = 3
    for r in prod:
        emit(ws, ri, r)
        ri += 1

    sub = ri
    ws.cell(row=sub, column=1, value="需生产 小计").font = bf
    ws.cell(row=sub, column=1).border = thin
    for col in (4, 5, 6, 7, 8, 10, 12, 13, 14, 15):
        lt = openpyxl.utils.get_column_letter(col)
        cell = ws.cell(row=sub, column=col, value=f"=SUM({lt}3:{lt}{sub-1})")
        cell.number_format = "#,##0"; cell.font = bf; cell.border = thin
    ri = sub + 2

    if noprod:
        ws.cell(row=ri, column=1, value="库存充足 / 无需生产").font = Font(bold=True, color="548235", size=11)
        ri += 1
        for r in noprod:
            emit(ws, ri, r)
            ri += 1

    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 42
    ws.freeze_panes = "C3"

    # ═══════════════════════════════════════════
    # Sheet 2: 验证推导
    # ═══════════════════════════════════════════
    ws2 = wb.create_sheet("验证推导")

    ws2.merge_cells("A1:L1")
    cell = ws2.cell(row=1, column=1,
        value=f"逐行验证: 库存 → 出货扣减 → 缺口 → 生产分配({P1} + {P2}) → 最终余额")
    cell.font = Font(bold=True, size=12, color="FFFFFF")
    cell.fill = F["banner"]
    cell.alignment = Alignment(horizontal="center")

    v_hdrs = [
        ("产品名称", 40, "blue"),
        ("客户型号", 18, "blue"),
        ("A: 总库存", 10, "blue"),
        ("B: 4/4前出货", 12, "gray"),
        ("C=A-B\n4/4后余额", 13, "gray"),
        ("D: 4/12前出货", 12, "gray"),
        ("E=C-D\n4/12后余额", 13, "gray"),
        ("F=max(-E,0)\n需要生产", 12, "orange"),
        (f"G: 生产\n{P1}", 12, "green"),
        (f"H: 生产\n{P2}", 12, "dblue"),
        (f"I=C+G\n4/4后(含{P1})", 15, "green"),
        ("J=E+G+H\n最终余额", 14, "dblue"),
    ]

    for col, (h, w, fk) in enumerate(v_hdrs, 1):
        cell = ws2.cell(row=2, column=col, value=h)
        cell.font = wf
        cell.fill = F[fk]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    ri = 3
    all_sorted = prod + noprod
    for r in all_sorted:
        inv = r["总库存"]
        w1 = r["4.4_出货个"]
        w2 = r["4.12_出货个"]
        bal1 = inv - w1
        bal2 = bal1 - w2
        p1 = r[f"生产_{P1}"]
        p2 = r[f"生产_{P2}"]
        final_w1 = bal1 + p1
        final_all = bal2 + p1 + p2

        vals = [
            r["产品名称"], r["客户型号"],
            inv, w1, bal1, w2, bal2,
            r["需要生产"], p1, p2,
            final_w1, final_all,
        ]

        for col, v in enumerate(vals, 1):
            cell = ws2.cell(row=ri, column=col, value=v)
            cell.border = thin
            if col >= 3:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")

            if col == 5 and bal1 < 0:
                cell.fill = F["lr"]; cell.font = rf
            elif col == 7 and bal2 < 0:
                cell.fill = F["lr"]; cell.font = rf
            elif col == 8 and r["需要生产"] > 0:
                cell.fill = F["lo"]
            elif col == 9 and p1 > 0:
                cell.fill = F["lg"]
            elif col == 10 and p2 > 0:
                cell.fill = F["lb"]
            elif col == 11:
                cell.fill = F["lg"] if final_w1 >= 0 else F["lr"]
                if final_w1 < 0: cell.font = rf
            elif col == 12:
                cell.fill = F["lb"] if final_all >= 0 else F["lr"]
                if final_all < 0: cell.font = rf
                cell.font = bf

        ri += 1

    ws2.cell(row=ri, column=1, value="合计").font = bf
    ws2.cell(row=ri, column=1).border = thin
    for col in range(3, 13):
        lt = openpyxl.utils.get_column_letter(col)
        cell = ws2.cell(row=ri, column=col, value=f"=SUM({lt}3:{lt}{ri-1})")
        cell.number_format = "#,##0"; cell.font = bf; cell.border = thin

    ws2.row_dimensions[1].height = 28
    ws2.row_dimensions[2].height = 48
    ws2.freeze_panes = "C3"

    # ═══════════════════════════════════════════
    # Sheet 3: 每日排产建议
    # ═══════════════════════════════════════════
    ws3 = wb.create_sheet("每日排产建议")
    s3_h = ["排产周期", "产品名称", "客户型号", "生产数量", "累计", "需天数"]
    for col, h in enumerate(s3_h, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = wf; cell.fill = F["dblue"]; cell.border = thin
        cell.alignment = Alignment(horizontal="center")

    plan = {
        f"{P1}\n(4/4前完成)": [(r["产品名称"], r["客户型号"], r[f"生产_{P1}"])
                              for r in prod if r[f"生产_{P1}"] > 0],
        f"{P2}\n(4/8前完成)": [(r["产品名称"], r["客户型号"], r[f"生产_{P2}"])
                              for r in prod if r[f"生产_{P2}"] > 0],
    }

    ri = 2
    for batch, items in plan.items():
        fill = F["lg"] if "4/1" in batch else F["lb"]
        cum = 0
        for name, sku, qty in items:
            cum += qty
            for col, v in enumerate([batch, name, sku, qty, cum, math.ceil(qty / DAILY_CAP)], 1):
                cell = ws3.cell(row=ri, column=col, value=v)
                cell.border = thin
                if col >= 4: cell.number_format = "#,##0"
                cell.fill = fill
            ri += 1
        wtot = sum(q for _, _, q in items)
        ws3.cell(row=ri, column=1, value=f"小计").font = bf
        ws3.cell(row=ri, column=1).border = thin
        for col, v in [(4, wtot), (6, math.ceil(wtot / DAILY_CAP))]:
            cell = ws3.cell(row=ri, column=col, value=v)
            cell.font = bf; cell.number_format = "#,##0"; cell.border = thin
        ri += 1

    for i, w in enumerate([14, 40, 18, 12, 12, 10], 1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws3.freeze_panes = "A2"

    OUT.parent.mkdir(exist_ok=True)
    wb.save(OUT)
    print(f"\n✅ {OUT}")


def main():
    rows = read_source()
    need_count = sum(1 for r in rows if r["需要生产"] > 0)
    total_need = sum(max(r["需要生产"], 0) for r in rows)
    print(f"{len(rows)} SKU, 需生产 {need_count} 个, 共 {total_need:,.0f}\n")
    rows = allocate(rows)
    write_excel(rows)


if __name__ == "__main__":
    main()
