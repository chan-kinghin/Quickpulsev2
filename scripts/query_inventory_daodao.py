#!/usr/bin/env python3
"""
Query 刀刀电子 sales orders (07.xx finished goods only),
group by (material_code, aux_prop_id) for true SKU uniqueness,
resolve aux_prop descriptions, enrich with per-warehouse inventory + undelivered.
"""

import json
import os
from collections import defaultdict
from pathlib import Path
from dotenv import load_dotenv
from k3cloud_webapi_sdk.main import K3CloudApiSdk

project_root = Path(__file__).parent.parent
load_dotenv(project_root / ".env")


def init_sdk():
    url = os.environ["KINGDEE_SERVER_URL"]
    sdk = K3CloudApiSdk(url)
    sdk.InitConfig(
        acct_id=os.environ["KINGDEE_ACCT_ID"],
        user_name=os.environ["KINGDEE_USER_NAME"],
        app_id=os.environ["KINGDEE_APP_ID"],
        app_secret=os.environ["KINGDEE_APP_SEC"],
        server_url=url,
        lcid=int(os.environ.get("KINGDEE_LCID", "2052")),
    )
    return sdk


def query_all_pages(sdk, form_id, fields, filter_str, order="", page_size=2000):
    all_rows = []
    start = 0
    while True:
        params = {
            "FormId": form_id,
            "FieldKeys": fields,
            "FilterString": filter_str,
            "OrderString": order,
            "Limit": page_size,
            "StartRow": start,
        }
        response = sdk.ExecuteBillQuery(params)
        result = json.loads(response)
        if not isinstance(result, list) or not result or isinstance(result[0], dict):
            if isinstance(result, list) and result and isinstance(result[0], dict):
                print(f"  ⚠️ API error for {form_id}: {json.dumps(result[0], ensure_ascii=False)[:200]}")
            break
        all_rows.extend(result)
        if len(result) < page_size:
            break
        start += page_size
    return all_rows


def main():
    sdk = init_sdk()

    # ══════════════════════════════════════════════════════════════
    # Step 1: Sales orders — 历史累计下单量
    # ══════════════════════════════════════════════════════════════
    print("① 查询 刀刀电子 销售订单 (07.xx 成品)...")
    so_fields = ",".join([
        "FBillNo", "FMaterialId.FNumber", "FMaterialId.FName",
        "FMaterialId.FSpecification", "FCustId.FName", "FQty",
        "FMtoNo", "FDate", "FAuxPropId",
    ])
    so_filter = (
        "FCustId.FName like '%刀刀%' "
        "and FDocumentStatus in ('B','C','D') "
        "and FMaterialId.FNumber like '07.%'"
    )
    so_rows = query_all_pages(sdk, "SAL_SaleOrder", so_fields, so_filter, "FDate DESC")
    print(f"  → {len(so_rows)} 条明细行\n")

    # Group by (material_code, aux_prop_id)
    sku_map = defaultdict(lambda: {
        "名称": "", "规格": "", "aux_prop_id": 0,
        "总销售量": 0, "订单数": 0, "最近MTO": "", "最近日期": "",
    })
    all_aux_ids = set()
    for row in so_rows:
        mat_num, aux_id = row[1], int(row[8]) if row[8] else 0
        key = (mat_num, aux_id)
        sku = sku_map[key]
        sku["名称"] = row[2]
        sku["规格"] = row[3]
        sku["aux_prop_id"] = aux_id
        sku["总销售量"] += row[5]
        sku["订单数"] += 1
        if not sku["最近日期"] or str(row[7]) > sku["最近日期"]:
            sku["最近日期"] = str(row[7])
            sku["最近MTO"] = str(row[6])
        if aux_id:
            all_aux_ids.add(aux_id)

    # ══════════════════════════════════════════════════════════════
    # Step 2: Undelivered orders — 未交订单量
    # ══════════════════════════════════════════════════════════════
    print("② 查询 刀刀电子 未交订单...")
    ud_fields = ",".join([
        "FBillNo", "FMaterialId.FNumber", "FAuxPropId", "FQty", "FStockOutQty",
    ])
    ud_filter = (
        "FCustId.FName like '%刀刀%' "
        "and FDocumentStatus in ('B','C','D') "
        "and FCloseStatus = 'A' "
        "and FMaterialId.FNumber like '07.%'"
    )
    ud_rows = query_all_pages(sdk, "SAL_SaleOrder", ud_fields, ud_filter)
    print(f"  → {len(ud_rows)} 条未交明细\n")

    undelivered = defaultdict(lambda: {"未交量": 0, "未交订单数": 0})
    for row in ud_rows:
        if len(row) < 5:
            continue
        mat_num, aux_id = row[1], int(row[2]) if row[2] else 0
        qty = row[3] if row[3] else 0
        out_qty = row[4] if row[4] else 0
        remain = qty - out_qty
        if remain > 0:
            key = (mat_num, aux_id)
            undelivered[key]["未交量"] += remain
            undelivered[key]["未交订单数"] += 1

    # ══════════════════════════════════════════════════════════════
    # Step 3: Inventory — per-warehouse breakdown
    # ══════════════════════════════════════════════════════════════
    print("③ 查询库存 (STK_Inventory) 按仓库明细...")
    mat_nums = set(k[0] for k in sku_map)
    mat_filter = " or ".join(f"FMaterialId.FNumber='{m}'" for m in mat_nums)
    inv_fields = "FMaterialId.FNumber,FStockId.FNumber,FStockId.FName,FAuxPropId,FBaseQty"
    inv_filter = f"FBaseQty <> 0 and ({mat_filter})"
    inv_rows = query_all_pages(sdk, "STK_Inventory", inv_fields, inv_filter)
    print(f"  → {len(inv_rows)} 条库存记录\n")

    # Discover all warehouse names + build per-SKU per-warehouse totals
    all_warehouses = set()
    # inv_detail[(mat, aux)][(wh_code, wh_name)] = qty
    inv_detail = defaultdict(lambda: defaultdict(float))
    for row in inv_rows:
        mat_num = row[0]
        wh_code, wh_name = row[1], row[2]
        aux_id = int(row[3]) if row[3] else 0
        qty = row[4] if row[4] else 0
        if aux_id:
            all_aux_ids.add(aux_id)
        wh_key = (wh_code, wh_name)
        all_warehouses.add(wh_key)
        inv_detail[(mat_num, aux_id)][wh_key] += qty

    # Sort warehouses: 成品仓 first, then others alphabetically
    def wh_sort_key(wh):
        code, name = wh
        if "成品" in name:
            return (0, code)
        return (1, code)

    sorted_warehouses = sorted(all_warehouses, key=wh_sort_key)
    print(f"  发现 {len(sorted_warehouses)} 个仓库:")
    for code, name in sorted_warehouses:
        print(f"    {code} — {name}")

    # ══════════════════════════════════════════════════════════════
    # Step 4: Resolve aux_prop descriptions
    # ══════════════════════════════════════════════════════════════
    print(f"\n④ 查询 {len(all_aux_ids)} 个辅助属性描述...")
    aux_descs = {}
    if all_aux_ids:
        id_list = list(all_aux_ids)
        for i in range(0, len(id_list), 200):
            batch = id_list[i:i + 200]
            in_clause = ",".join(str(x) for x in batch)
            aux_rows = query_all_pages(
                sdk, "BD_FLEXSITEMDETAILV",
                "FID,FF100001,FF100002.FName",
                f"FID IN ({in_clause})",
            )
            for r in aux_rows:
                fid = int(r[0])
                spec = str(r[1]).strip() if r[1] else ""
                color = str(r[2]).strip() if r[2] else ""
                aux_descs[fid] = spec or color or ""
    print(f"  → 解析到 {len(aux_descs)} 个描述\n")

    # ══════════════════════════════════════════════════════════════
    # Step 5: Export to Excel
    # ══════════════════════════════════════════════════════════════
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        print("⚠️ openpyxl not installed")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "刀刀电子_成品SKU"

    # ── Build headers ──
    # Fixed columns
    fixed_headers = [
        "物料编码", "物料名称", "规格型号", "辅助属性ID", "辅助属性描述",
        "累计订单数", "累计销售量",
        "未交订单数", "未交量",
    ]
    # One column per warehouse
    wh_headers = [f"{name}\n({code})" for code, name in sorted_warehouses]
    # Summary columns after warehouses
    tail_headers = ["库存合计", "最近MTO", "最近日期"]

    all_headers = fixed_headers + wh_headers + tail_headers

    # Styles
    blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    green_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    orange_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    gray_fill = PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    light_green = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    light_orange = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True, size=10)
    red_font = Font(color="FF0000", bold=True)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Write header row
    wh_start_col = len(fixed_headers) + 1
    wh_end_col = wh_start_col + len(sorted_warehouses) - 1
    tail_start_col = wh_end_col + 1

    for col, h in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
        # Color code header sections
        if col <= 5:
            cell.fill = blue_fill  # SKU info
        elif col <= 7:
            cell.fill = gray_fill  # 累计销售
        elif col <= 9:
            cell.fill = orange_fill  # 未交
        elif col <= wh_end_col:
            cell.fill = green_fill  # 仓库
        else:
            cell.fill = blue_fill  # tail

    # Write data rows
    row_idx = 2
    for (mat_num, aux_id), sku in sorted(sku_map.items()):
        aux_desc = aux_descs.get(aux_id, "") if aux_id else ""
        ud = undelivered.get((mat_num, aux_id), {})
        uq = ud.get("未交量", 0)
        udn = ud.get("未交订单数", 0)
        wh_data = inv_detail.get((mat_num, aux_id), {})

        # Fixed columns
        fixed = [
            mat_num, sku["名称"], sku["规格"], aux_id, aux_desc,
            sku["订单数"], sku["总销售量"],
            udn, uq,
        ]

        # Per-warehouse quantities
        wh_qtys = []
        total_inv = 0
        for wh_key in sorted_warehouses:
            q = wh_data.get(wh_key, 0)
            wh_qtys.append(q if q != 0 else None)  # None = leave cell empty for clarity
            total_inv += q

        # Tail
        tail = [total_inv, sku["最近MTO"], sku["最近日期"][:10]]

        row_data = fixed + wh_qtys + tail

        for col, val in enumerate(row_data, 1):
            if val is None:
                continue  # leave empty
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin

            # Number formatting for quantity columns
            if col in (6, 7, 8, 9) or wh_start_col <= col <= wh_end_col or col == tail_start_col:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")

            # Background tinting per section
            if col in (8, 9):
                cell.fill = light_orange
            elif wh_start_col <= col <= wh_end_col:
                cell.fill = light_green

        # Highlight risk: 未交 > 0 but 库存合计 = 0
        if uq > 0 and total_inv == 0:
            for c in range(wh_start_col, tail_start_col + 1):
                cell = ws.cell(row=row_idx, column=c)
                cell.fill = yellow_fill
                cell.font = red_font

        row_idx += 1

    # ── Totals row ──
    total_row = row_idx
    ws.cell(row=total_row, column=1, value="合计").font = Font(bold=True, size=11)
    ws.cell(row=total_row, column=1).border = thin
    # Sum formulas for numeric columns
    for col in range(6, len(all_headers) + 1):
        # Skip non-numeric tail columns (MTO, date)
        if col >= tail_start_col + 1:
            continue
        from_letter = openpyxl.utils.get_column_letter(col)
        cell = ws.cell(
            row=total_row, column=col,
            value=f"=SUM({from_letter}2:{from_letter}{total_row - 1})",
        )
        cell.number_format = "#,##0"
        cell.font = Font(bold=True)
        cell.border = thin
        cell.alignment = Alignment(horizontal="right")

    # ── Column widths ──
    fixed_widths = [16, 14, 28, 12, 32, 10, 12, 10, 10]
    for i, w in enumerate(fixed_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    for i in range(wh_start_col, wh_end_col + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 14
    ws.column_dimensions[openpyxl.utils.get_column_letter(tail_start_col)].width = 12
    ws.column_dimensions[openpyxl.utils.get_column_letter(tail_start_col + 1)].width = 16
    ws.column_dimensions[openpyxl.utils.get_column_letter(tail_start_col + 2)].width = 12

    # Row height for header
    ws.row_dimensions[1].height = 40

    # Freeze panes + filter
    ws.freeze_panes = "F2"
    ws.auto_filter.ref = ws.dimensions

    out_path = project_root / "reports" / "刀刀电子_成品SKU汇总.xlsx"
    out_path.parent.mkdir(exist_ok=True)
    wb.save(out_path)
    print(f"\n✅ Excel saved to: {out_path}")

    # Summary
    print(f"\n{'='*60}")
    print(f"SKU 数: {len(sku_map)}")
    print(f"仓库数: {len(sorted_warehouses)}")
    print(f"Excel 列数: {len(all_headers)}")
    print(f"  固定列: {len(fixed_headers)} (SKU信息 + 销售 + 未交)")
    print(f"  仓库列: {len(sorted_warehouses)} (每个仓库一列)")
    print(f"  尾部列: {len(tail_headers)} (库存合计 + MTO + 日期)")
    print(f"仓库清单:")
    for code, name in sorted_warehouses:
        print(f"  {code} — {name}")


if __name__ == "__main__":
    main()
