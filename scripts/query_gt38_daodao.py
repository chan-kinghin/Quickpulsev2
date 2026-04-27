#!/usr/bin/env python3
"""
Extract all GT38 and variants for 刀刀电子 — across all material levels
(07.xx 成品, 06.xx 未包装, 05.xx 半成品, etc.)
"""

import json
import os
from collections import defaultdict
from pathlib import Path
from dotenv import load_dotenv
from k3cloud_webapi_sdk.main import K3CloudApiSdk

project_root = Path(__file__).parent.parent
load_dotenv(project_root / ".env")


def init_sdk():
    url = os.environ["KINGDEE_SERVER_URL"]
    sdk = K3CloudApiSdk(url)
    sdk.InitConfig(
        acct_id=os.environ["KINGDEE_ACCT_ID"],
        user_name=os.environ["KINGDEE_USER_NAME"],
        app_id=os.environ["KINGDEE_APP_ID"],
        app_secret=os.environ["KINGDEE_APP_SEC"],
        server_url=url,
        lcid=int(os.environ.get("KINGDEE_LCID", "2052")),
    )
    return sdk


def query_all(sdk, form_id, fields, filter_str, order="", page_size=2000):
    all_rows = []
    start = 0
    while True:
        params = {
            "FormId": form_id, "FieldKeys": fields,
            "FilterString": filter_str, "OrderString": order,
            "Limit": page_size, "StartRow": start,
        }
        resp = json.loads(sdk.ExecuteBillQuery(params))
        if not isinstance(resp, list) or not resp or isinstance(resp[0], dict):
            break
        all_rows.extend(resp)
        if len(resp) < page_size:
            break
        start += page_size
    return all_rows


def main():
    sdk = init_sdk()

    # ═══════════════════════════════════════════════════
    # 1. Sales orders — 刀刀 + GT38 (all material levels)
    # ═══════════════════════════════════════════════════
    print("① 查询 刀刀电子 × GT38 销售订单 (所有物料级别)...")
    so_fields = ",".join([
        "FBillNo", "FMaterialId.FNumber", "FMaterialId.FName",
        "FMaterialId.FSpecification", "FCustId.FName", "FQty",
        "FMtoNo", "FDate", "FAuxPropId",
    ])
    so_filter = (
        "FCustId.FName like '%刀刀%' "
        "and FDocumentStatus in ('B','C','D') "
        "and FMaterialId.FSpecification like '%GT38%'"
    )
    so_rows = query_all(sdk, "SAL_SaleOrder", so_fields, so_filter, "FDate DESC")
    print(f"  → {len(so_rows)} 条\n")

    # Group by (mat, aux)
    sku_map = defaultdict(lambda: {
        "名称": "", "规格": "", "aux_prop_id": 0,
        "总销售量": 0, "订单数": 0, "最近MTO": "", "最近日期": "",
        "订单明细": [],
    })
    all_aux_ids = set()
    for r in so_rows:
        mat, aux = r[1], int(r[8]) if r[8] else 0
        k = (mat, aux)
        s = sku_map[k]
        s["名称"], s["规格"], s["aux_prop_id"] = r[2], r[3], aux
        s["总销售量"] += r[5]
        s["订单数"] += 1
        s["订单明细"].append({
            "订单号": r[0], "数量": r[5], "MTO": r[6], "日期": str(r[7])[:10],
        })
        if not s["最近日期"] or str(r[7]) > s["最近日期"]:
            s["最近日期"] = str(r[7])
            s["最近MTO"] = str(r[6])
        if aux:
            all_aux_ids.add(aux)

    print(f"  SKU 数: {len(sku_map)} (物料 {len(set(k[0] for k in sku_map))} 种)\n")

    # ═══════════════════════════════════════════════════
    # 2. Undelivered
    # ═══════════════════════════════════════════════════
    print("② 查询未交订单...")
    ud_fields = "FBillNo,FMaterialId.FNumber,FAuxPropId,FQty,FStockOutQty,FDate"
    ud_filter = (
        "FCustId.FName like '%刀刀%' "
        "and FDocumentStatus in ('B','C','D') "
        "and FCloseStatus = 'A' "
        "and FMaterialId.FSpecification like '%GT38%'"
    )
    ud_rows = query_all(sdk, "SAL_SaleOrder", ud_fields, ud_filter)
    print(f"  → {len(ud_rows)} 条\n")

    undelivered = defaultdict(lambda: {"未交量": 0, "未交订单数": 0, "明细": []})
    for r in ud_rows:
        if len(r) < 5:
            continue
        mat, aux = r[1], int(r[2]) if r[2] else 0
        remain = (r[3] or 0) - (r[4] or 0)
        if remain > 0:
            k = (mat, aux)
            undelivered[k]["未交量"] += remain
            undelivered[k]["未交订单数"] += 1
            undelivered[k]["明细"].append({
                "订单号": r[0], "订单量": r[3], "已出库": r[4], "未交": remain,
                "日期": str(r[5])[:10] if len(r) > 5 else "",
            })

    # ═══════════════════════════════════════════════════
    # 3. Inventory — all materials with GT38 in spec
    # ═══════════════════════════════════════════════════
    print("③ 查询库存...")
    # Query all materials that appeared in sales, plus broader GT38 search
    mat_nums = set(k[0] for k in sku_map)
    mat_filter = " or ".join(f"FMaterialId.FNumber='{m}'" for m in mat_nums)
    inv_fields = "FMaterialId.FNumber,FMaterialId.FName,FStockId.FNumber,FStockId.FName,FAuxPropId,FBaseQty"
    inv_filter = f"FBaseQty <> 0 and ({mat_filter})"
    inv_rows = query_all(sdk, "STK_Inventory", inv_fields, inv_filter)
    print(f"  → {len(inv_rows)} 条\n")

    all_warehouses = set()
    inv_detail = defaultdict(lambda: defaultdict(float))
    for r in inv_rows:
        mat, wh_code, wh_name = r[0], r[2], r[3]
        aux = int(r[4]) if r[4] else 0
        qty = r[5] if r[5] else 0
        if aux:
            all_aux_ids.add(aux)
        wh = (wh_code, wh_name)
        all_warehouses.add(wh)
        inv_detail[(mat, aux)][wh] += qty

    def wh_sort(w):
        return (0 if "成品" in w[1] else 1, w[0])

    sorted_wh = sorted(all_warehouses, key=wh_sort)
    print(f"  仓库: {[n for _, n in sorted_wh]}\n")

    # ═══════════════════════════════════════════════════
    # 4. Aux prop descriptions
    # ═══════════════════════════════════════════════════
    print(f"④ 查询 {len(all_aux_ids)} 个辅助属性...")
    aux_descs = {}
    if all_aux_ids:
        for i in range(0, len(list(all_aux_ids)), 200):
            batch = list(all_aux_ids)[i:i + 200]
            in_cl = ",".join(str(x) for x in batch)
            for r in query_all(sdk, "BD_FLEXSITEMDETAILV", "FID,FF100001,FF100002.FName", f"FID IN ({in_cl})"):
                fid = int(r[0])
                spec = str(r[1]).strip() if r[1] else ""
                color = str(r[2]).strip() if r[2] else ""
                aux_descs[fid] = spec or color or ""
    print(f"  → {len(aux_descs)} 个\n")

    # ═══════════════════════════════════════════════════
    # 5. Excel — two sheets
    # ═══════════════════════════════════════════════════
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    white_bold = Font(color="FFFFFF", bold=True, size=10)
    bold = Font(bold=True)

    fills = {
        "blue": PatternFill("solid", fgColor="4472C4"),
        "gray": PatternFill("solid", fgColor="A5A5A5"),
        "orange": PatternFill("solid", fgColor="ED7D31"),
        "green": PatternFill("solid", fgColor="548235"),
        "light_orange": PatternFill("solid", fgColor="FCE4D6"),
        "light_green": PatternFill("solid", fgColor="E2EFDA"),
        "yellow": PatternFill("solid", fgColor="FFF2CC"),
        "red_font": Font(color="FF0000", bold=True),
    }

    # ────── Sheet 1: SKU 汇总 ──────
    ws = wb.active
    ws.title = "GT38_SKU汇总"

    fixed_h = [
        "物料编码", "物料名称", "规格型号", "辅助属性ID", "辅助属性描述",
        "累计订单数", "累计销售量", "未交订单数", "未交量",
    ]
    wh_h = [f"{n}\n({c})" for c, n in sorted_wh]
    tail_h = ["库存合计", "最近MTO", "最近日期"]
    headers = fixed_h + wh_h + tail_h

    wh_start = len(fixed_h) + 1
    wh_end = wh_start + len(sorted_wh) - 1
    tail_start = wh_end + 1

    # Header colors by section
    section_fills = {}
    for i in range(1, 6):
        section_fills[i] = fills["blue"]
    for i in range(6, 8):
        section_fills[i] = fills["gray"]
    for i in range(8, 10):
        section_fills[i] = fills["orange"]
    for i in range(wh_start, wh_end + 1):
        section_fills[i] = fills["green"]
    for i in range(tail_start, tail_start + 3):
        section_fills[i] = fills["blue"]

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = white_bold
        c.fill = section_fills.get(col, fills["blue"])
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin

    row_idx = 2
    for (mat, aux), sku in sorted(sku_map.items()):
        desc = aux_descs.get(aux, "") if aux else ""
        ud = undelivered.get((mat, aux), {})
        uq, udn = ud.get("未交量", 0), ud.get("未交订单数", 0)
        wh_data = inv_detail.get((mat, aux), {})

        vals = [
            mat, sku["名称"], sku["规格"], aux, desc,
            sku["订单数"], sku["总销售量"], udn, uq,
        ]
        total_inv = 0
        for wk in sorted_wh:
            q = wh_data.get(wk, 0)
            vals.append(q if q else None)
            total_inv += q
        vals += [total_inv, sku["最近MTO"], sku["最近日期"][:10]]

        for col, v in enumerate(vals, 1):
            if v is None:
                continue
            c = ws.cell(row=row_idx, column=col, value=v)
            c.border = thin
            if col in (6, 7, 8, 9) or wh_start <= col <= wh_end or col == tail_start:
                c.number_format = "#,##0"
                c.alignment = Alignment(horizontal="right")
            if col in (8, 9):
                c.fill = fills["light_orange"]
            elif wh_start <= col <= wh_end:
                c.fill = fills["light_green"]

        # Highlight: undelivered > 0 but no inventory
        if uq > 0 and total_inv == 0:
            for cc in range(wh_start, tail_start + 1):
                cell = ws.cell(row=row_idx, column=cc)
                cell.fill = fills["yellow"]
                cell.font = fills["red_font"]

        row_idx += 1

    # Totals row
    tr = row_idx
    ws.cell(row=tr, column=1, value="合计").font = bold
    ws.cell(row=tr, column=1).border = thin
    for col in range(6, tail_start + 1):
        letter = openpyxl.utils.get_column_letter(col)
        c = ws.cell(row=tr, column=col, value=f"=SUM({letter}2:{letter}{tr - 1})")
        c.number_format = "#,##0"
        c.font = bold
        c.border = thin
        c.alignment = Alignment(horizontal="right")

    # Widths
    for i, w in enumerate([16, 14, 22, 12, 34, 10, 12, 10, 10], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    for i in range(wh_start, wh_end + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 14
    for i, w in zip(range(tail_start, tail_start + 3), [12, 16, 12]):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 40
    ws.freeze_panes = "F2"
    ws.auto_filter.ref = ws.dimensions

    # ────── Sheet 2: 未交订单明细 ──────
    ws2 = wb.create_sheet("未交订单明细")
    ud_headers = ["物料编码", "辅助属性描述", "订单号", "订单量", "已出库", "未交量", "日期"]
    for col, h in enumerate(ud_headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = white_bold
        c.fill = fills["orange"]
        c.alignment = Alignment(horizontal="center")
        c.border = thin

    r = 2
    for (mat, aux), ud in sorted(undelivered.items()):
        desc = aux_descs.get(aux, "") if aux else ""
        for d in sorted(ud["明细"], key=lambda x: x["日期"], reverse=True):
            for col, v in enumerate([mat, desc, d["订单号"], d["订单量"], d["已出库"], d["未交"], d["日期"]], 1):
                c = ws2.cell(row=r, column=col, value=v)
                c.border = thin
                if col in (4, 5, 6):
                    c.number_format = "#,##0"
            r += 1

    for i, w in enumerate([16, 34, 16, 10, 10, 10, 12], 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions

    # ────── Sheet 3: 销售订单明细 ──────
    ws3 = wb.create_sheet("历史订单明细")
    so_headers = ["物料编码", "辅助属性描述", "订单号", "数量", "MTO号", "日期"]
    for col, h in enumerate(so_headers, 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font = white_bold
        c.fill = fills["gray"]
        c.alignment = Alignment(horizontal="center")
        c.border = thin

    r = 2
    for (mat, aux), sku in sorted(sku_map.items()):
        desc = aux_descs.get(aux, "") if aux else ""
        for d in sorted(sku["订单明细"], key=lambda x: x["日期"], reverse=True):
            for col, v in enumerate([mat, desc, d["订单号"], d["数量"], d["MTO"], d["日期"]], 1):
                c = ws3.cell(row=r, column=col, value=v)
                c.border = thin
                if col == 4:
                    c.number_format = "#,##0"
            r += 1

    for i, w in enumerate([16, 34, 16, 10, 16, 12], 1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = ws3.dimensions

    # Save
    out = project_root / "reports" / "刀刀电子_GT38汇总.xlsx"
    out.parent.mkdir(exist_ok=True)
    wb.save(out)
    print(f"✅ Excel saved to: {out}")
    print(f"  Sheet 1: GT38_SKU汇总 ({len(sku_map)} SKUs × {len(headers)} 列)")
    print(f"  Sheet 2: 未交订单明细 ({sum(len(u['明细']) for u in undelivered.values())} 行)")
    print(f"  Sheet 3: 历史订单明细 ({sum(len(s['订单明细']) for s in sku_map.values())} 行)")


if __name__ == "__main__":
    main()
